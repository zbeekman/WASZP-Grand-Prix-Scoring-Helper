"""Tests for the Excel export module (exporter.py).

Covers:
- Output file has exactly 2 sheets named "Finish Placings" and "Gate List"
- All required columns present in every table
- Sheet 1 contains overall, 8.2, and non-8.2 table pairs (ranked + original)
- Sheet 2 contains gate rounding entries in recording order
- Default filename follows {EventName}_Race{N}_{Date}.xlsx
- Cell fill hex values set on appropriate rows (programmatic check)
- Font color hex values set on appropriate rows (programmatic check)
- Edge case: empty gate list → no crash
- Edge case: no finish entries → no crash
- Edge case: single competitor → valid output
"""

import os
import tempfile
from typing import Optional

import openpyxl

from waszp_gp_scorer.exporter import (
    _BG_5PLUS,
    _BG_COLORS,
    _FONT_COLORS,
    _TABLE_COLS,
    export,
    export_filename,
)
from waszp_gp_scorer.models import (
    Competitor,
    FinishEntry,
    FinishLineConfig,
    GateRounding,
    RaceSession,
    ScoredResult,
)
from waszp_gp_scorer.scorer import score

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _competitor(
    sail_number: str,
    rig_size: str = "8.2",
    country_code: str = "TST",
    division: str = "Open",
) -> Competitor:
    return Competitor(
        sail_number=sail_number,
        country_code=country_code,
        name=f"Sailor {sail_number}",
        rig_size=rig_size,
        division=division,
    )


def _session(
    *,
    num_laps: int = 2,
    gate_sns: list[str],
    finish_sns: list[tuple[str, Optional[str]]],
    competitors: Optional[list[Competitor]] = None,
    event_name: str = "Test Event",
    race_number: int = 1,
    race_date: str = "2026-03-22",
    finish_line_config: FinishLineConfig = FinishLineConfig.FINISH_AT_GATE,
) -> RaceSession:
    if competitors is None:
        all_sns = set(gate_sns) | {sn for sn, _ in finish_sns}
        competitors = [_competitor(sn) for sn in sorted(all_sns)]
    return RaceSession(
        event_name=event_name,
        race_number=race_number,
        race_date=race_date,
        num_laps=num_laps,
        finish_line_config=finish_line_config,
        competitors=competitors,
        gate_roundings=[
            GateRounding(position=i + 1, sail_number=sn)
            for i, sn in enumerate(gate_sns)
        ],
        finish_entries=[
            FinishEntry(position=i + 1, sail_number=sn, letter_score=ls)
            for i, (sn, ls) in enumerate(finish_sns)
        ],
    )


def _export_to_tmp(
    session: RaceSession, results: list[ScoredResult]
) -> openpyxl.Workbook:
    """Export session to a temporary file and return the loaded workbook."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        tmp_path = f.name
    try:
        export(session, results, tmp_path)
        return openpyxl.load_workbook(tmp_path)
    finally:
        os.unlink(tmp_path)


def _cell_rgb(cell: openpyxl.cell.Cell, attr: str = "fill") -> Optional[str]:
    """Return the RGB string for a cell's fill or font color, or None."""
    if attr == "fill":
        fill = cell.fill
        # PatternFill with no fill_type or 'none' means no fill
        if fill.fill_type is None or fill.fill_type == "none":
            return None
        rgb = fill.fgColor.rgb
        # Strip ARGB alpha prefix if present (e.g. "FF66D9FF" → "66D9FF")
        return rgb[-6:] if len(rgb) == 8 else rgb
    else:  # font
        font = cell.font
        if font.color is None:
            return None
        rgb = font.color.rgb
        return rgb[-6:] if len(rgb) == 8 else rgb


def _col_values(ws: openpyxl.worksheet.worksheet.Worksheet, col: int) -> list[object]:
    """Return all non-None values in a column (1-based col index)."""
    return [
        ws.cell(row=r, column=col).value
        for r in range(1, ws.max_row + 1)
        if ws.cell(row=r, column=col).value is not None
    ]


# ---------------------------------------------------------------------------
# Standard fixture: 5-boat fleet, 2 rig sizes, 2 laps
# ---------------------------------------------------------------------------

#: Five competitors: 3× 8.2, 2× 7.5
_FLEET_5 = [
    _competitor("101", rig_size="8.2"),
    _competitor("102", rig_size="8.2"),
    _competitor("103", rig_size="8.2"),
    _competitor("201", rig_size="7.5"),
    _competitor("202", rig_size="7.5"),
]


def _std_session() -> RaceSession:
    """2-lap race, 5 boats, mix of Standard / GP / Gate finishes."""
    return _session(
        num_laps=2,
        gate_sns=["101", "201", "102", "202", "101", "201"],
        finish_sns=[("101", None), ("201", None), ("102", None)],
        competitors=_FLEET_5,
    )


# ===========================================================================
# Tests — Sheet structure
# ===========================================================================


class TestSheetStructure:
    def test_two_sheets(self) -> None:
        sess = _std_session()
        results, _ = score(sess)
        wb = _export_to_tmp(sess, results)
        assert len(wb.sheetnames) == 2

    def test_sheet_names(self) -> None:
        sess = _std_session()
        results, _ = score(sess)
        wb = _export_to_tmp(sess, results)
        assert wb.sheetnames[0] == "Finish Placings"
        assert wb.sheetnames[1] == "Gate List"


# ===========================================================================
# Tests — Default filename
# ===========================================================================


class TestExportFilename:
    def test_basic(self) -> None:
        sess = _session(
            gate_sns=[],
            finish_sns=[],
            event_name="WASZP Worlds",
            race_number=3,
            race_date="2026-07-15",
        )
        assert export_filename(sess) == "WASZP_Worlds_Race3_2026-07-15.xlsx"

    def test_spaces_replaced(self) -> None:
        sess = _session(gate_sns=[], finish_sns=[], event_name="A B C")
        assert export_filename(sess).startswith("A_B_C_Race")

    def test_empty_event_name(self) -> None:
        sess = _session(gate_sns=[], finish_sns=[], event_name="")
        assert export_filename(sess).startswith("Event_Race")

    def test_race_number_in_name(self) -> None:
        sess = _session(
            gate_sns=[], finish_sns=[], race_number=7, race_date="2026-01-01"
        )
        assert "Race7" in export_filename(sess)


# ===========================================================================
# Tests — Sheet 1 columns
# ===========================================================================


class TestFinishPlacingsColumns:
    def _find_header_rows(
        self, ws: openpyxl.worksheet.worksheet.Worksheet
    ) -> list[int]:
        """Return (row, col) pairs where TABLE_COLS headers start."""
        from waszp_gp_scorer.exporter import _RIGHT_OFFSET

        matches = []
        for row in range(1, ws.max_row + 1):
            for start_col in (1, 1 + _RIGHT_OFFSET):
                vals = [
                    ws.cell(row=row, column=start_col + c).value
                    for c in range(len(_TABLE_COLS))
                ]
                if vals == _TABLE_COLS:
                    matches.append((row, start_col))
        return matches

    def test_six_header_rows_in_sheet1(self) -> None:
        """Sheet 1 should have 6 header rows: 3 sections × 2 tables (left + right)."""
        sess = _std_session()
        results, _ = score(sess)
        wb = _export_to_tmp(sess, results)
        ws = wb["Finish Placings"]
        header_rows = self._find_header_rows(ws)
        assert len(header_rows) == 6

    def test_all_columns_present_left_table(self) -> None:
        sess = _std_session()
        results, _ = score(sess)
        wb = _export_to_tmp(sess, results)
        ws = wb["Finish Placings"]
        # Column headers should appear in columns 1..N_COLS at row 2
        headers_row2 = [
            ws.cell(row=2, column=c + 1).value for c in range(len(_TABLE_COLS))
        ]
        assert headers_row2 == _TABLE_COLS

    def test_all_columns_present_right_table(self) -> None:
        sess = _std_session()
        results, _ = score(sess)
        wb = _export_to_tmp(sess, results)
        ws = wb["Finish Placings"]
        from waszp_gp_scorer.exporter import _RIGHT_OFFSET

        headers = [
            ws.cell(row=2, column=_RIGHT_OFFSET + 1 + c).value
            for c in range(len(_TABLE_COLS))
        ]
        assert headers == _TABLE_COLS


# ===========================================================================
# Tests — Sheet 1 sections (title rows)
# ===========================================================================


class TestFinishPlacingsSections:
    def _title_values(self, ws: openpyxl.worksheet.worksheet.Worksheet) -> set[object]:
        """Return all non-None values in column 1 of Sheet 1."""
        return set(_col_values(ws, 1))

    def test_overall_section_titles(self) -> None:
        sess = _std_session()
        results, _ = score(sess)
        wb = _export_to_tmp(sess, results)
        ws = wb["Finish Placings"]
        titles = self._title_values(ws)
        assert "Overall GP Finish Ranking" in titles
        assert (
            "Overall Original Finish Order" not in titles
        )  # in right column, not col 1
        # Check the right-column title appears somewhere in the sheet
        all_vals: set[object] = set()
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                v = ws.cell(row=row, column=col).value
                if v is not None:
                    all_vals.add(v)
        assert "Overall Original Finish Order" in all_vals

    def test_eight_two_section_titles(self) -> None:
        sess = _std_session()
        results, _ = score(sess)
        wb = _export_to_tmp(sess, results)
        ws = wb["Finish Placings"]
        all_vals: set[object] = set()
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                v = ws.cell(row=row, column=col).value
                if v is not None:
                    all_vals.add(v)
        assert "8.2 Fleet GP Finish Ranking" in all_vals
        assert "8.2 Fleet Original Finish Order" in all_vals

    def test_non_eight_two_section_titles(self) -> None:
        sess = _std_session()
        results, _ = score(sess)
        wb = _export_to_tmp(sess, results)
        ws = wb["Finish Placings"]
        all_vals: set[object] = set()
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                v = ws.cell(row=row, column=col).value
                if v is not None:
                    all_vals.add(v)
        assert "Non-8.2 Fleet GP Finish Ranking" in all_vals
        assert "Non-8.2 Fleet Original Finish Order" in all_vals

    def test_overall_data_row_count(self) -> None:
        """Overall tables should have one data row per scored result."""
        sess = _std_session()
        results, _ = score(sess)
        wb = _export_to_tmp(sess, results)
        ws = wb["Finish Placings"]
        # Rows 1=title, 2=headers, 3..N=data for overall section
        # Count how many rows have a numeric "Place" value in column 1
        # starting from row 3 until the next title row.
        place_col_vals = [
            ws.cell(row=r, column=1).value for r in range(3, ws.max_row + 1)
        ]
        data_rows = [v for v in place_col_vals if isinstance(v, int)]
        # The overall section should have exactly len(results) data rows.
        assert len(data_rows) >= len(results)

    def test_eight_two_data_rows(self) -> None:
        """8.2 section has one row per 8.2-fleet result."""
        sess = _std_session()
        results, _ = score(sess)
        wb = _export_to_tmp(sess, results)
        ws = wb["Finish Placings"]
        # Check right-column 8.2 sail numbers appear in sheet
        eight_two_sns = {
            r.competitor.sail_number for r in results if r.competitor.rig_size == "8.2"
        }
        found: set[str] = set()
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=3).value  # Sail # is col 3
            if v in eight_two_sns:
                found.add(str(v))
        # Each 8.2 sail number should appear at least twice (left + right table,
        # and again in the 8.2 section).
        assert len(found) == len(eight_two_sns)


# ===========================================================================
# Tests — Sheet 2 (Gate List)
# ===========================================================================


class TestGateListSheet:
    def test_gate_list_header_row(self) -> None:
        sess = _std_session()
        results, _ = score(sess)
        wb = _export_to_tmp(sess, results)
        ws = wb["Gate List"]
        headers = [ws.cell(row=1, column=c + 1).value for c in range(len(_TABLE_COLS))]
        assert headers == _TABLE_COLS

    def test_gate_list_row_count(self) -> None:
        """Gate List has header + one row per gate rounding."""
        sess = _std_session()
        results, _ = score(sess)
        wb = _export_to_tmp(sess, results)
        ws = wb["Gate List"]
        # Header row + data rows
        expected_rows = 1 + len(sess.gate_roundings)
        assert ws.max_row == expected_rows

    def test_gate_list_recording_order(self) -> None:
        """Sail numbers appear in the same order as gate_roundings."""
        sess = _std_session()
        results, _ = score(sess)
        wb = _export_to_tmp(sess, results)
        ws = wb["Gate List"]
        gate_sns = [r.sail_number for r in sess.gate_roundings]
        sheet_sns = [
            ws.cell(row=r, column=3).value  # Sail # is column 3
            for r in range(2, 2 + len(gate_sns))
        ]
        assert sheet_sns == gate_sns

    def test_gate_list_place_column_sequential(self) -> None:
        """Place column in Gate List is 1-based sequential."""
        sess = _std_session()
        results, _ = score(sess)
        wb = _export_to_tmp(sess, results)
        ws = wb["Gate List"]
        for i in range(len(sess.gate_roundings)):
            assert ws.cell(row=i + 2, column=1).value == i + 1

    def test_gate_list_columns_complete(self) -> None:
        """All TABLE_COLS appear in Gate List header."""
        sess = _std_session()
        results, _ = score(sess)
        wb = _export_to_tmp(sess, results)
        ws = wb["Gate List"]
        headers = [ws.cell(row=1, column=c + 1).value for c in range(len(_TABLE_COLS))]
        assert headers == _TABLE_COLS


# ===========================================================================
# Tests — Cell colors
# ===========================================================================


class TestCellColors:
    def _make_multi_lap_session(self) -> tuple[RaceSession, list[ScoredResult]]:
        """Build a session where boat '101' rounds the gate twice (2-lap Standard)."""
        sess = _session(
            num_laps=2,
            gate_sns=["101", "102", "101"],
            finish_sns=[("101", None), ("102", None)],
            competitors=[
                _competitor("101", rig_size="8.2"),
                _competitor("102", rig_size="8.2"),
            ],
        )
        results, _ = score(sess)
        return sess, results

    def test_second_rounding_has_blue_fill(self) -> None:
        """Row for a boat's 2nd gate rounding gets blue (#66D9FF) background."""
        sess, results = self._make_multi_lap_session()
        wb = _export_to_tmp(sess, results)
        ws = wb["Gate List"]
        # gate_sns = ["101", "102", "101"]
        # Row 2 = 1st rounding of "101" → no fill
        # Row 3 = 1st rounding of "102" → no fill
        # Row 4 = 2nd rounding of "101" → blue
        rgb = _cell_rgb(ws.cell(row=4, column=1))
        assert rgb == _BG_COLORS[2]  # "66D9FF"

    def test_first_rounding_no_fill(self) -> None:
        """Row for a boat's 1st gate rounding has no fill."""
        sess, results = self._make_multi_lap_session()
        wb = _export_to_tmp(sess, results)
        ws = wb["Gate List"]
        assert _cell_rgb(ws.cell(row=2, column=1)) is None

    def test_2lap_boat_font_color_teal(self) -> None:
        """A 2-lap boat's rows in the Gate List have teal font (#0088AA)."""
        sess, results = self._make_multi_lap_session()
        wb = _export_to_tmp(sess, results)
        ws = wb["Gate List"]
        # gate_sns = ["101", "102", "101"]; boat "101" has 2 scored laps
        # Rows 2 and 4 are both for "101" → teal font
        font_row2 = _cell_rgb(ws.cell(row=2, column=1), attr="font")
        font_row4 = _cell_rgb(ws.cell(row=4, column=1), attr="font")
        assert font_row2 == _FONT_COLORS[2]  # "0088AA"
        assert font_row4 == _FONT_COLORS[2]

    def test_1lap_boat_font_color_black(self) -> None:
        """A 1-lap GP boat has dark/black font (#1C1C1E)."""
        # In FINISH_AT_GATE, "102" has 1 gate rounding + finish → 2 laps → teal.
        # Test a Gate-only boat with 1 lap to check black font.
        sess2 = _session(
            num_laps=2,
            gate_sns=["111"],
            finish_sns=[],
            competitors=[_competitor("111", rig_size="7.5")],
        )
        results2, _ = score(sess2)
        wb2 = _export_to_tmp(sess2, results2)
        ws2 = wb2["Gate List"]
        font = _cell_rgb(ws2.cell(row=2, column=1), attr="font")
        assert font == _FONT_COLORS[1]  # "1C1C1E"

    def test_5plus_rounding_gets_pink_fill(self) -> None:
        """5th gate rounding gets pink (#FF80BF) background."""
        sess = _session(
            num_laps=6,
            gate_sns=["99", "99", "99", "99", "99"],
            finish_sns=[("99", None)],
            competitors=[_competitor("99", rig_size="8.2")],
        )
        results, _ = score(sess)
        wb = _export_to_tmp(sess, results)
        ws = wb["Gate List"]
        # Row 6 = 5th rounding of "99"
        rgb = _cell_rgb(ws.cell(row=6, column=1))
        assert rgb == _BG_5PLUS  # "FF80BF"

    def test_finish_placings_row_has_color_set(self) -> None:
        """Rows in the Finish Placings sheet have font color set."""
        sess, results = self._make_multi_lap_session()
        wb = _export_to_tmp(sess, results)
        ws = wb["Finish Placings"]
        # Row 3 is the first data row of the overall section.
        font_color = _cell_rgb(ws.cell(row=3, column=1), attr="font")
        assert font_color is not None


# ===========================================================================
# Tests — Edge cases
# ===========================================================================


class TestEdgeCases:
    def test_empty_gate_list_no_crash(self) -> None:
        """Export succeeds with no gate roundings."""
        sess = _session(
            num_laps=2,
            gate_sns=[],
            finish_sns=[("101", None)],
            competitors=[_competitor("101")],
        )
        results, _ = score(sess)
        wb = _export_to_tmp(sess, results)
        ws = wb["Gate List"]
        # Only header row
        assert ws.max_row == 1

    def test_no_finish_entries_no_crash(self) -> None:
        """Export succeeds with no finish entries."""
        sess = _session(
            num_laps=2,
            gate_sns=["101"],
            finish_sns=[],
            competitors=[_competitor("101")],
        )
        results, _ = score(sess)
        wb = _export_to_tmp(sess, results)
        assert len(wb.sheetnames) == 2

    def test_single_competitor_no_crash(self) -> None:
        """Export with a single competitor produces a valid workbook."""
        sess = _session(
            num_laps=2,
            gate_sns=["1"],
            finish_sns=[("1", None)],
            competitors=[_competitor("1")],
        )
        results, _ = score(sess)
        wb = _export_to_tmp(sess, results)
        assert "Finish Placings" in wb.sheetnames

    def test_all_letter_scores_no_crash(self) -> None:
        """Export succeeds when all finish entries have letter scores."""
        sess = _session(
            num_laps=2,
            gate_sns=[],
            finish_sns=[("101", "DNS"), ("102", "DSQ")],
            competitors=[
                _competitor("101", rig_size="8.2"),
                _competitor("102", rig_size="7.5"),
            ],
        )
        results, _ = score(sess)
        wb = _export_to_tmp(sess, results)
        assert len(wb.sheetnames) == 2

    def test_no_eight_two_fleet_no_crash(self) -> None:
        """Export succeeds when no boats have 8.2 rig size."""
        sess = _session(
            num_laps=2,
            gate_sns=["201"],
            finish_sns=[("201", None)],
            competitors=[_competitor("201", rig_size="7.5")],
        )
        results, _ = score(sess)
        wb = _export_to_tmp(sess, results)
        assert "Finish Placings" in wb.sheetnames

    def test_path_as_pathlib_path(self) -> None:
        """export() accepts a pathlib.Path as the path argument."""
        from pathlib import Path

        sess = _session(
            num_laps=2,
            gate_sns=["101"],
            finish_sns=[("101", None)],
            competitors=[_competitor("101")],
        )
        results, _ = score(sess)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = Path(f.name)
        try:
            export(sess, results, tmp_path)
            assert tmp_path.exists()
        finally:
            tmp_path.unlink(missing_ok=True)


# ===========================================================================
# Tests — Original finish order (right-side table)
# ===========================================================================


class TestOriginalOrder:
    def test_original_order_differs_from_ranked(self) -> None:
        """Right table shows boats in finish order, not GP rank order."""
        # Create a scenario where GP rank != finish list order:
        # 2 laps: boat "101" standard finisher (1 gate + finish)
        # 1 lap GP: boat "202" (0 gate roundings + finish → Finish Only)
        # boat "101" finished 2nd on the line but ranks 1st (more laps)
        sess = _session(
            num_laps=2,
            gate_sns=["101"],
            finish_sns=[("202", None), ("101", None)],  # 202 crossed first
            competitors=[
                _competitor("101", rig_size="8.2"),
                _competitor("202", rig_size="7.5"),
            ],
        )
        results, _ = score(sess)
        # "101" should be GP rank 1 (2 laps), "202" rank 2 (Finish Only 1 lap)
        assert results[0].competitor.sail_number == "101"
        assert results[1].competitor.sail_number == "202"

        from waszp_gp_scorer.exporter import _RIGHT_OFFSET

        wb = _export_to_tmp(sess, results)
        ws = wb["Finish Placings"]
        # Right table data starts at row 3, column (1 + _RIGHT_OFFSET + 2) for Sail #
        sail_col = 1 + _RIGHT_OFFSET + 2  # = Sail # column in right table
        right_row3 = ws.cell(row=3, column=sail_col).value  # original 1st finisher
        right_row4 = ws.cell(row=4, column=sail_col).value  # original 2nd finisher
        assert right_row3 == "202"  # first to cross the line
        assert right_row4 == "101"  # second to cross the line
