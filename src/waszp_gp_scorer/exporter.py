"""Excel export module for the WASZP GP Scorer application.

Produces a two-sheet ``.xlsx`` file from a scored race session.

Public API::

    export_filename(session) -> str
    export(session, scored_results, path)
"""

from __future__ import annotations

from pathlib import Path
from typing import Optional, Union

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from waszp_gp_scorer.models import RaceSession, ScoredResult

# ---------------------------------------------------------------------------
# Color constants (openpyxl RGB hex, 6-char, no '#' prefix)
# ---------------------------------------------------------------------------

#: Background fill hex values indexed by rounding number (1-based).
#: Rounding 1 has no fill; rounding 5+ uses :data:`_BG_5PLUS`.
_BG_COLORS: dict[int, str] = {
    2: "66D9FF",
    3: "CCFF66",
    4: "FFFF66",
}
_BG_5PLUS: str = "FF80BF"

#: Font color hex values indexed by total scored laps.
#: 5+ laps uses :data:`_FONT_5PLUS`.
_FONT_COLORS: dict[int, str] = {
    1: "1C1C1E",
    2: "0088AA",
    3: "5A9900",
    4: "9E9E00",
}
_FONT_5PLUS: str = "AA0055"

#: Column names shared by every results table.
_TABLE_COLS: list[str] = [
    "Place",
    "Country",
    "Sail #",
    "Sailor Name",
    "Rig Size",
    "Division",
    "Laps",
    "Finish Type",
]

_N_COLS: int = len(_TABLE_COLS)
#: Number of empty columns between the left and right table of a pair.
_GAP: int = 1
#: Column offset for the right-side table within a table pair.
_RIGHT_OFFSET: int = _N_COLS + _GAP


# ---------------------------------------------------------------------------
# Public helpers
# ---------------------------------------------------------------------------


def export_filename(session: RaceSession) -> str:
    """Return the default export filename for *session*.

    Format: ``{EventName}_Race{N}_{Date}.xlsx``

    Spaces in the event name are replaced with underscores.  Placeholder
    text is used if ``event_name`` or ``race_date`` are empty.

    Args:
        session: The race session.

    Returns:
        A filename string (not a full path).
    """
    event = session.event_name.replace(" ", "_") if session.event_name else "Event"
    date = session.race_date if session.race_date else "0000-00-00"
    return f"{event}_Race{session.race_number}_{date}.xlsx"


# ---------------------------------------------------------------------------
# Public export function
# ---------------------------------------------------------------------------


def export(
    session: RaceSession,
    scored_results: list[ScoredResult],
    path: Union[str, Path],
) -> None:
    """Export a scored race to a two-sheet ``.xlsx`` file.

    **Sheet 1 — Finish Placings**: three table pairs (overall, 8.2 fleet,
    non-8.2 fleet), each with a GP-ranked left table and an
    original-finish-order right table.

    **Sheet 2 — Gate List**: every gate rounding entry in recording order,
    with background fill and font colors reflecting the rounding tier and
    the boat's total scored laps.

    Args:
        session: The race session (competitor list, gate roundings, etc.).
        scored_results: Ordered list of
            :class:`~waszp_gp_scorer.models.ScoredResult` from
            :func:`~waszp_gp_scorer.scorer.score`.
        path: Filesystem path at which to save the ``.xlsx`` file.
    """
    wb = Workbook()
    ws1: Worksheet = wb.active
    ws1.title = "Finish Placings"
    _write_finish_placings(ws1, session, scored_results)

    ws2: Worksheet = wb.create_sheet("Gate List")
    _write_gate_list(ws2, session, scored_results)

    wb.save(str(path))


# ---------------------------------------------------------------------------
# Internal — color helpers
# ---------------------------------------------------------------------------


def _bg_fill(tier: int) -> Optional[PatternFill]:
    """Return a solid :class:`~openpyxl.styles.PatternFill` for *tier*, or ``None``.

    *tier* is the rounding/lap number (1-based).  Tier 1 returns ``None``
    (no fill — white background).
    """
    if tier <= 1:
        return None
    hex_color = _BG_COLORS.get(tier, _BG_5PLUS)
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _font_for(laps: int) -> Font:
    """Return a :class:`~openpyxl.styles.Font` for *laps* total scored laps."""
    color = _FONT_COLORS.get(laps, _FONT_5PLUS)
    return Font(color=color)


def _apply_row_style(
    ws: Worksheet,
    row: int,
    start_col: int,
    n_cols: int,
    laps: int,
    rounding_tier: Optional[int] = None,
) -> None:
    """Apply background fill and font color to a range of cells in *row*.

    Args:
        ws: The worksheet.
        row: 1-based row index.
        start_col: 1-based starting column index.
        n_cols: Number of columns to style.
        laps: Scored laps — determines font color.
        rounding_tier: If provided, overrides *laps* for background color
            selection (used for gate list rows where background reflects the
            rounding number rather than the scored laps).
    """
    tier = rounding_tier if rounding_tier is not None else laps
    fill = _bg_fill(tier)
    font = _font_for(laps)
    for col in range(start_col, start_col + n_cols):
        cell = ws.cell(row=row, column=col)
        if fill is not None:
            cell.fill = fill
        cell.font = font


# ---------------------------------------------------------------------------
# Internal — table writing helpers
# ---------------------------------------------------------------------------


def _write_headers(ws: Worksheet, row: int, col: int) -> None:
    """Write column header row starting at (*row*, *col*)."""
    for i, name in enumerate(_TABLE_COLS):
        ws.cell(row=row, column=col + i).value = name


def _write_result_row(
    ws: Worksheet,
    row: int,
    col: int,
    result: ScoredResult,
    place: Union[int, str, None] = None,
) -> None:
    """Write a single :class:`~waszp_gp_scorer.models.ScoredResult` to *row*.

    Args:
        ws: Target worksheet.
        row: 1-based row index.
        col: 1-based starting column index.
        result: The scored result to write.
        place: Override for the Place cell.  Defaults to ``result.place``.
    """
    comp = result.competitor
    place_val: Union[int, str] = place if place is not None else result.place
    finish_type_str = (
        result.letter_score
        if result.letter_score is not None
        else result.finish_type.value
    )
    values: list[Union[int, str]] = [
        place_val,
        comp.country_code,
        comp.sail_number,
        comp.name,
        comp.rig_size,
        comp.division,
        result.laps,
        finish_type_str,
    ]
    for i, val in enumerate(values):
        ws.cell(row=row, column=col + i).value = val
    _apply_row_style(ws, row, col, _N_COLS, result.laps)


def _original_order(
    session: RaceSession,
    results: list[ScoredResult],
) -> list[ScoredResult]:
    """Return *results* sorted by original finish list order.

    Boats that appear on the finish list come first, sorted by their last
    recorded finish list position.  Boats not on the finish list (gate-only)
    appear at the end in their original GP-rank order.

    Args:
        session: Race session (provides ``finish_entries``).
        results: Scored results to reorder.

    Returns:
        Reordered list of scored results.
    """
    result_by_sail = {r.competitor.sail_number: r for r in results}

    # Last occurrence wins (mirrors scorer duplicate handling).
    finish_pos: dict[str, int] = {}
    for fe in session.finish_entries:
        if fe.sail_number in result_by_sail:
            finish_pos[fe.sail_number] = fe.position

    finishers: list[ScoredResult] = [
        result_by_sail[sn]
        for sn, _pos in sorted(finish_pos.items(), key=lambda kv: kv[1])
    ]
    gate_only: list[ScoredResult] = [
        r for r in results if r.competitor.sail_number not in finish_pos
    ]
    return finishers + gate_only


def _filter_fleet(
    results: list[ScoredResult],
    eight_two: bool,
) -> list[ScoredResult]:
    """Return results for the 8.2 fleet or the non-8.2 fleet.

    Args:
        results: Full scored results list.
        eight_two: If ``True``, return only 8.2-rig boats; otherwise return
            all non-8.2-rig boats.

    Returns:
        Filtered list preserving input order.
    """
    if eight_two:
        return [r for r in results if r.competitor.rig_size == "8.2"]
    return [r for r in results if r.competitor.rig_size != "8.2"]


def _write_table_pair(
    ws: Worksheet,
    start_row: int,
    ranked: list[ScoredResult],
    original: list[ScoredResult],
    ranked_title: str,
    original_title: str,
) -> int:
    """Write a GP-ranked (left) + original-order (right) table pair.

    Layout (columns)::

        1..N_COLS | gap | (N_COLS+GAP+1)..(2*N_COLS+GAP)
        [ranked]  |     | [original order]

    Args:
        ws: Target worksheet.
        start_row: First row to write to.
        ranked: Results in GP-ranked order (for the left table).
        original: Results in original finish order (for the right table).
        ranked_title: Section title for the left table.
        original_title: Section title for the right table.

    Returns:
        Next available row number after this pair (including a blank separator).
    """
    # Section title row
    ws.cell(row=start_row, column=1).value = ranked_title
    ws.cell(row=start_row, column=1 + _RIGHT_OFFSET).value = original_title

    # Column headers
    _write_headers(ws, start_row + 1, 1)
    _write_headers(ws, start_row + 1, 1 + _RIGHT_OFFSET)

    data_start = start_row + 2

    # Left table — GP ranked
    for offset, result in enumerate(ranked):
        _write_result_row(ws, data_start + offset, 1, result)

    # Right table — original finish order (sequential place 1, 2, 3…)
    for offset, result in enumerate(original):
        _write_result_row(
            ws,
            data_start + offset,
            1 + _RIGHT_OFFSET,
            result,
            place=offset + 1,
        )

    n_rows = max(len(ranked), len(original), 0)
    # +1 for blank separator row between sections
    return data_start + n_rows + 1


# ---------------------------------------------------------------------------
# Internal — sheet writers
# ---------------------------------------------------------------------------


def _write_finish_placings(
    ws: Worksheet,
    session: RaceSession,
    scored_results: list[ScoredResult],
) -> None:
    """Populate the **Finish Placings** worksheet with three table pairs.

    Sections (top to bottom):
    1. Overall (all competitors)
    2. 8.2 fleet only
    3. Non-8.2 fleet only

    Args:
        ws: Target worksheet (already titled).
        session: Race session.
        scored_results: GP-ranked scored results.
    """
    overall_orig = _original_order(session, scored_results)
    row = _write_table_pair(
        ws,
        start_row=1,
        ranked=scored_results,
        original=overall_orig,
        ranked_title="Overall GP Finish Ranking",
        original_title="Overall Original Finish Order",
    )

    eight_two = _filter_fleet(scored_results, eight_two=True)
    eight_two_orig = _original_order(session, eight_two)
    row = _write_table_pair(
        ws,
        start_row=row,
        ranked=eight_two,
        original=eight_two_orig,
        ranked_title="8.2 Fleet GP Finish Ranking",
        original_title="8.2 Fleet Original Finish Order",
    )

    non_eight_two = _filter_fleet(scored_results, eight_two=False)
    non_eight_two_orig = _original_order(session, non_eight_two)
    _write_table_pair(
        ws,
        start_row=row,
        ranked=non_eight_two,
        original=non_eight_two_orig,
        ranked_title="Non-8.2 Fleet GP Finish Ranking",
        original_title="Non-8.2 Fleet Original Finish Order",
    )


def _write_gate_list(
    ws: Worksheet,
    session: RaceSession,
    scored_results: list[ScoredResult],
) -> None:
    """Populate the **Gate List** worksheet.

    Each gate rounding entry occupies one row, in recording order.
    Colors applied per row:

    - **Background**: based on the rounding tier for that boat (which
      rounding of the race this entry represents — 1st = no fill, 2nd =
      blue, 3rd = green, etc.).
    - **Font**: based on the boat's total scored laps (from *scored_results*,
      falling back to raw gate rounding count if the boat has no scored
      result).

    Args:
        ws: Target worksheet (already titled).
        session: Race session providing gate roundings and competitor list.
        scored_results: Scored results used to look up final lap counts.
    """
    competitor_map = {c.sail_number: c for c in session.competitors}
    result_map = {r.competitor.sail_number: r for r in scored_results}

    _write_headers(ws, 1, 1)

    # Running rounding count per boat (for background color tier).
    boat_rounding_count: dict[str, int] = {}
    # Pre-compute total gate roundings per boat (fallback font color).
    total_gate_roundings: dict[str, int] = {}
    for rounding in session.gate_roundings:
        sn = rounding.sail_number
        total_gate_roundings[sn] = total_gate_roundings.get(sn, 0) + 1

    for seq_idx, rounding in enumerate(session.gate_roundings):
        sn = rounding.sail_number
        boat_rounding_count[sn] = boat_rounding_count.get(sn, 0) + 1
        rounding_tier = boat_rounding_count[sn]

        comp = competitor_map.get(sn)
        result = result_map.get(sn)

        country = comp.country_code if comp else "UNK"
        name = comp.name if comp else f"Unknown ({sn})"
        rig_size = comp.rig_size if comp else "Unknown"
        division = comp.division if comp else "Unknown"
        # Use scored laps for Laps column; fall back to raw gate count.
        laps = result.laps if result is not None else total_gate_roundings[sn]
        finish_type_str: str
        if result is not None:
            finish_type_str = (
                result.letter_score
                if result.letter_score is not None
                else result.finish_type.value
            )
        else:
            finish_type_str = ""

        row = seq_idx + 2  # header is row 1; data starts at row 2
        values: list[Union[int, str]] = [
            seq_idx + 1,  # Place = 1-based sequence position
            country,
            sn,
            name,
            rig_size,
            division,
            laps,
            finish_type_str,
        ]
        for col_idx, val in enumerate(values):
            ws.cell(row=row, column=col_idx + 1).value = val

        # Background: rounding tier; font: scored laps.
        _apply_row_style(
            ws,
            row=row,
            start_col=1,
            n_cols=_N_COLS,
            laps=laps,
            rounding_tier=rounding_tier,
        )
